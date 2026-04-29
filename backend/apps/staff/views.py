import secrets
import string
import datetime
from collections import defaultdict
from django.db import transaction
from django.db import models
from django.http import HttpResponse
from rest_framework import permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from .models import Staff, TeacherDetail, StaffAttendance, StaffLeaveRequest, StaffTask, TeacherLeaderboardSnapshot, LeavePolicy, LeaveBalance
from .serializers import StaffSerializer, TeacherDetailSerializer, StaffCreateSerializer, StaffUpdateSerializer, StaffAttendanceSerializer, StaffLeaveRequestSerializer, StaffTaskSerializer, TeacherLeaderboardSnapshotSerializer
from django.utils import timezone
from apps.accounts.models import User
from apps.permissions.models import Role as RoleV2
from apps.academics.models import Subject

def _random_password(length=12):
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))

STAFF_LATE_CUTOFF = datetime.time(9, 0)

def _late_attendance_q():
    # Backward compatible: some records may store `status='late'` without setting `is_late`.
    return models.Q(is_late=True) | models.Q(late_minutes__gt=0) | models.Q(status='late')

def _to_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(value)
    except (TypeError, ValueError):
        return default

def _to_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("1", "true", "t", "yes", "y", "on"):
        return True
    if s in ("0", "false", "f", "no", "n", "off"):
        return False
    return None

def _compute_late_fields(selected_date, in_time, late_minutes, force_late=False):
    late_minutes = max(0, _to_int(late_minutes, 0))

    if force_late:
        return True, late_minutes

    if late_minutes > 0:
        return True, late_minutes

    if not in_time:
        return False, 0

    if in_time > STAFF_LATE_CUTOFF:
        delta = datetime.datetime.combine(selected_date, in_time) - datetime.datetime.combine(selected_date, STAFF_LATE_CUTOFF)
        return True, int(delta.total_seconds() // 60)

    return False, 0


def _get_staff_department(staff):
    """Dynamically map staff to departments based on their roles if department is 'General'."""
    if staff.department and staff.department != 'General':
        return staff.department
    
    role_name = staff.user.role.name if staff.user.role else 'Unassigned'
    
    mapping = {
        'Academic': ['Academic', 'Teacher', 'Subject Teacher', 'Class Teacher', 'Student'],
        'Administration': ['Admin', 'Super Admin', 'Principal', 'Staff'],
        'Finance & Accounts': ['Accounts', 'Accountant'],
        'Student Support & Welfare': ['Guidance Counselor', 'Warden'],
        'Operations & Support': ['Driver', 'Support Staff', 'Unassigned']
    }
    
    for dept, roles in mapping.items():
        if any(role.lower() == role_name.lower() for role in roles):
            return dept
            
    return 'General'


DEPARTMENT_ROLE_MAP = {
    'Academic': ['Academic', 'Teacher', 'Subject Teacher', 'Class Teacher', 'Student'],
    'Administration': ['Admin', 'Super Admin', 'Principal', 'Staff'],
    'Finance & Accounts': ['Accounts', 'Accountant'],
    'Student Support & Welfare': ['Guidance Counselor', 'Warden'],
    'Operations & Support': ['Driver', 'Support Staff', 'Unassigned'],
}


def _filter_staff_by_department(staff_qs, department):
    """Filter a Staff queryset by logical department name."""
    if not department:
        return staff_qs
    roles = DEPARTMENT_ROLE_MAP.get(department)
    if roles:
        return staff_qs.filter(user__role__name__iregex=f"^({'|'.join(roles)})$")
    # Fallback: try direct department field match
    return staff_qs.filter(
        models.Q(department__iexact=department) |
        models.Q(user__role__name__iexact=department)
    )


@api_view(['GET', 'POST'])
@permission_classes([permissions.IsAuthenticated])
def staff_list_view(request):
    """List all staff with optional role filtering."""
    if request.method == "GET":
        q = request.query_params.get('q')
        status_filter = request.query_params.get('status')
        role_filter = request.query_params.get('role')  # teacher | non_teacher
        role_id = request.query_params.get('role_id')
        qs = Staff.objects.select_related("user", "user__role", "teacher_detail").all()

        if role_filter == 'teacher':
            qs = qs.filter(is_teaching_staff=True)
        elif role_filter == 'non_teacher':
            qs = qs.filter(is_teaching_staff=False)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if role_id:
            qs = qs.filter(user__role_id=role_id)
        if q:
            qs = qs.filter(
                models.Q(employee_id__icontains=q) |
                models.Q(user__first_name__icontains=q) |
                models.Q(user__last_name__icontains=q) |
                models.Q(user__email__icontains=q) |
                models.Q(user__phone__icontains=q)
            )

        serializer = StaffSerializer(qs, many=True)
        return Response(serializer.data)

    serializer = StaffCreateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    data = serializer.validated_data

    username = data.get("username") or (data.get("email") or "").split("@")[0] or f"staff_{data['employee_id']}"
    email = data.get("email") or ""
    phone = data.get("phone") or ""
    role_id = data.get("role")
    role = RoleV2.objects.filter(pk=role_id).first() if role_id else None
    requested_subject_ids = data.get("teaching_subject_ids", []) or []

    with transaction.atomic():
        user = User.objects.create(
            username=username,
            first_name=data["first_name"],
            last_name=data.get("last_name") or "",
            email=email,
            phone=phone,
            portal="admin",
            school=request.user.school,
            role=role,
        )
        user.set_password(data["password"])
        user.save()

        # Handle auto-generation of employee_id if not provided
        emp_id = data.get("employee_id")
        if not emp_id:
            emp_id = Staff.generate_next_id()

        staff = Staff.objects.create(
            user=user,
            employee_id=emp_id,
            designation=data["designation"],
            joining_date=data["joining_date"],
            status=data.get("status") or "active",
            is_teaching_staff=data.get("is_teaching_staff", True),
        )
        if staff.is_teaching_staff:
            teacher_detail, _ = TeacherDetail.objects.get_or_create(
                staff=staff,
                defaults={
                    "specialization": data.get("specialization", ""),
                    "bio": data.get("bio", ""),
                },
            )
            if requested_subject_ids:
                allowed_subjects = Subject.objects.filter(id__in=requested_subject_ids).filter(
                    models.Q(school=request.user.school) | models.Q(school__isnull=True)
                )
                teacher_detail.teaching_subjects.set(allowed_subjects)

    return Response(StaffSerializer(staff).data, status=status.HTTP_201_CREATED)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def teacher_list_view(request):
    """Detailed teacher list with academic allocations."""
    qs = Staff.objects.select_related("user", "user__role", "teacher_detail").filter(is_teaching_staff=True)
    serializer = StaffSerializer(qs, many=True)
    
    # Enrich with teacher detail if needed
    return Response(serializer.data)

@api_view(['GET', 'PATCH'])
@permission_classes([permissions.IsAuthenticated])
def staff_detail_view(request, pk):
    try:
        staff = Staff.objects.select_related("user", "user__role", "teacher_detail").get(pk=pk)
    except Staff.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
        
    if request.method == 'GET':
        serializer = StaffSerializer(staff)
        # Include teacher details if applicable
        data = serializer.data
        if staff.is_teaching_staff and hasattr(staff, 'teacher_detail'):
            data['teacher_info'] = TeacherDetailSerializer(staff.teacher_detail).data
        return Response(data)
        
    elif request.method == 'PATCH':
        serializer = StaffUpdateSerializer(data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        data = serializer.validated_data

        if "status" in data:
            staff.status = data["status"]
        if "designation" in data:
            staff.designation = data["designation"]
        if "is_teaching_staff" in data:
            staff.is_teaching_staff = data["is_teaching_staff"]

        if "role" in data:
            staff.user.role_id = data["role"]
        if "is_active" in data:
            staff.user.is_active = data["is_active"]
        if "teaching_subject_ids" in data and staff.is_teaching_staff:
            teacher_detail, _ = TeacherDetail.objects.get_or_create(staff=staff)
            requested_subject_ids = data.get("teaching_subject_ids", []) or []
            allowed_subjects = Subject.objects.filter(id__in=requested_subject_ids).filter(
                models.Q(school=staff.user.school) | models.Q(school__isnull=True)
            )
            teacher_detail.teaching_subjects.set(allowed_subjects)
        staff.user.save()
        staff.save()
        return Response(StaffSerializer(staff).data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def staff_reset_credentials_view(request, pk):
    try:
        staff = Staff.objects.select_related("user").get(pk=pk)
    except Staff.DoesNotExist:
        return Response({"error": "Staff not found."}, status=status.HTTP_404_NOT_FOUND)

    temp_password = _random_password(12)
    staff.user.set_password(temp_password)
    staff.user.save(update_fields=["password"])
    return Response({"username": staff.user.username, "temp_password": temp_password}, status=status.HTTP_200_OK)


@api_view(['GET', 'POST'])
@permission_classes([permissions.IsAuthenticated])
def staff_attendance_view(request):
    if request.method == 'GET':
        qs = StaffAttendance.objects.all()
        serializer = StaffAttendanceSerializer(qs, many=True)
        return Response(serializer.data)
    
    # POST - Clock In / Clock Out
    action = request.data.get('action')
    try:
        staff = request.user.staff_profile
    except:
        return Response({"error": "Not a staff member"}, status=status.HTTP_403_FORBIDDEN)
        
    date_today = timezone.now().date()
    attendance, created = StaffAttendance.objects.get_or_create(staff=staff, date=date_today)
    
    if action == 'clock_in':
        if attendance.in_time:
            return Response({"error": "Already clocked in"}, status=status.HTTP_400_BAD_REQUEST)
        attendance.in_time = timezone.now().time()
        attendance.status = 'present'
        # Simple late logic: if after 9:00 AM
        if attendance.in_time.hour >= 9:
            attendance.is_late = True
        attendance.save()
    elif action == 'clock_out':
        if not attendance.in_time:
            return Response({"error": "Clock in first"}, status=status.HTTP_400_BAD_REQUEST)
        attendance.out_time = timezone.now().time()
        attendance.save()
    else:
        return Response({"error": "Invalid action"}, status=status.HTTP_400_BAD_REQUEST)
        
    serializer = StaffAttendanceSerializer(attendance)
    return Response(serializer.data)

@api_view(['GET', 'POST'])
@permission_classes([permissions.IsAuthenticated])
def staff_leave_view(request):
    if request.method == 'GET':
        qs = StaffLeaveRequest.objects.all()
        serializer = StaffLeaveRequestSerializer(qs, many=True)
        return Response(serializer.data)
        
    # Apply for leave
    try:
        staff = request.user.staff_profile
    except:
        return Response({"error": "Not a staff member"}, status=status.HTTP_403_FORBIDDEN)
        
    data = request.data.copy()
    data['staff'] = staff.id
    serializer = StaffLeaveRequestSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PATCH'])
@permission_classes([permissions.IsAuthenticated])
def staff_leave_detail_view(request, pk):
    try:
        leave = StaffLeaveRequest.objects.get(pk=pk)
    except StaffLeaveRequest.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
        
    data = request.data
    if 'status' in data:
        leave.status = data['status']
        leave.reviewed_by = request.user
        leave.reviewed_at = timezone.now()
        leave.save()
        
        # If approved, auto mark attendance
        if leave.status == 'approved':
            import datetime
            current_date = leave.from_date
            while current_date <= leave.to_date:
                StaffAttendance.objects.update_or_create(
                    staff=leave.staff,
                    date=current_date,
                    defaults={'status': 'on_leave'}
                )
                current_date += datetime.timedelta(days=1)
                
    return Response(StaffLeaveRequestSerializer(leave).data)

@api_view(['GET', 'POST'])
@permission_classes([permissions.IsAuthenticated])
def staff_task_view(request):
    if request.method == 'GET':
        qs = StaffTask.objects.all()
        serializer = StaffTaskSerializer(qs, many=True)
        return Response(serializer.data)
        
    serializer = StaffTaskSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(assigned_by=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PATCH'])
@permission_classes([permissions.IsAuthenticated])
def staff_task_detail_view(request, pk):
    try:
        task = StaffTask.objects.get(pk=pk)
    except StaffTask.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    serializer = StaffTaskSerializer(task, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def teacher_leaderboard_view(request):
    qs = TeacherLeaderboardSnapshot.objects.all().order_by('-composite_score')
    serializer = TeacherLeaderboardSnapshotSerializer(qs, many=True)
    return Response(serializer.data)


# ─── HR Dashboard ───────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def staff_hr_dashboard_view(request):
    """Returns dashboard stats, trends, alerts for a selected date."""
    try:
        date_param = request.query_params.get('date')
        selected_date = datetime.datetime.strptime(date_param, '%Y-%m-%d').date() if date_param else timezone.now().date()
        month = selected_date.month
        year = selected_date.year

        staff_qs = Staff.objects.select_related('user').filter(status='active')
        total_staff = staff_qs.count()

        # Today's attendance
        today_att = StaffAttendance.objects.filter(date=selected_date).select_related('staff__user')
        present_today = today_att.filter(status='present', is_late=False).count()
        absent_today = today_att.filter(status='absent').count()
        late_today = today_att.filter(_late_attendance_q()).count()
        leave_today = today_att.filter(status='on_leave').count()
        half_day_today = today_att.filter(status='half_day').count()
        unmarked_today = max(0, total_staff - today_att.count())

        # Previous day for trend
        prev_date = selected_date - datetime.timedelta(days=1)
        prev_late = StaffAttendance.objects.filter(date=prev_date).filter(_late_attendance_q()).count()
        late_trend = late_today - prev_late

        # Weekly trend (last 7 days)
        weekly_trend = []
        for i in range(6, -1, -1):
            d = selected_date - datetime.timedelta(days=i)
            day_att = StaffAttendance.objects.filter(date=d)
            weekly_trend.append({
                'date': d.strftime('%Y-%m-%d'),
                'day': d.strftime('%a'),
                'present': day_att.filter(status='present', is_late=False).count(),
                'absent': day_att.filter(status='absent').count(),
                'late': day_att.filter(_late_attendance_q()).count(),
                'leave': day_att.filter(status='on_leave').count(),
            })

        # Monthly trend
        monthly_trend = []
        for w in range(1, 5):
            start_d = datetime.date(year, month, (w - 1) * 7 + 1) if (w - 1) * 7 + 1 <= 31 else datetime.date(year, month, 1)
            end_d = datetime.date(year, month, min(w * 7, 31)) if w * 7 <= 31 else datetime.date(year, month, 28)
            week_att = StaffAttendance.objects.filter(date__gte=start_d, date__lte=end_d, date__month=month, date__year=year)
            monthly_trend.append({
                'week': f'W{w}',
                'present': week_att.filter(status='present', is_late=False).count(),
                'absent': week_att.filter(status='absent').count(),
                'late': week_att.filter(_late_attendance_q()).count(),
            })

        # Department breakdown
        dept_map = {k: {'total': 0, 'present': 0, 'absent': 0, 'late': 0, 'leave': 0} for k in DEPARTMENT_ROLE_MAP.keys()}
        for s in staff_qs:
            dept = _get_staff_department(s)
            if dept not in dept_map:
                dept_map[dept] = {'total': 0, 'present': 0, 'absent': 0, 'late': 0, 'leave': 0}
            dept_map[dept]['total'] += 1

        for rec in today_att:
            dept = _get_staff_department(rec.staff)
            if rec.status == 'present' and not rec.is_late:
                dept_map[dept]['present'] += 1
            elif rec.status == 'absent':
                dept_map[dept]['absent'] += 1
            elif rec.status == 'on_leave':
                dept_map[dept]['leave'] += 1
            elif rec.status == 'late' or rec.is_late:
                dept_map[dept]['late'] += 1

        # Most frequent late employees (current month)
        late_employees = []
        month_late = StaffAttendance.objects.filter(date__month=month, date__year=year).filter(_late_attendance_q()).values('staff').annotate(count=models.Count('id')).order_by('-count')[:5]
        for item in month_late:
            stf = Staff.objects.select_related('user').filter(pk=item['staff']).first()
            if stf:
                late_employees.append({'id': stf.id, 'name': stf.user.get_full_name() or stf.user.username, 'late_count': item['count']})

        # Low attendance alert (<75%)
        low_attendance_staff = []
        for s in staff_qs:
            month_records = StaffAttendance.objects.filter(staff=s, date__month=month, date__year=year)
            total_days = month_records.count()
            if total_days > 0:
                present_days = month_records.filter(status__in=['present', 'half_day', 'late']).count()
                pct = round((present_days / total_days) * 100, 2)
                if pct < 75:
                    low_attendance_staff.append({'id': s.id, 'name': s.user.get_full_name() or s.user.username, 'percentage': pct, 'department': _get_staff_department(s)})

        # Peak absence days (current month)
        absence_days = StaffAttendance.objects.filter(date__month=month, date__year=year, status='absent').values('date').annotate(count=models.Count('id')).order_by('-count')[:5]
        peak_absence = [{'date': str(a['date']), 'count': a['count']} for a in absence_days]

        return Response({
            'date': selected_date.strftime('%Y-%m-%d'),
            'summary': {
                'total_staff': total_staff,
                'present_today': present_today,
                'absent_today': absent_today,
                'late_today': late_today,
                'leave_today': leave_today,
                'half_day_today': half_day_today,
                'unmarked_today': unmarked_today,
                'present_pct': round(((present_today + late_today + half_day_today) / total_staff) * 100, 1) if total_staff else 0,
                'late_trend': late_trend,
            },
            'weekly_trend': weekly_trend,
            'monthly_trend': monthly_trend,
            'department_breakdown': [{'department': k, **v} for k, v in dept_map.items()],
            'frequent_late_employees': late_employees,
            'low_attendance_alerts': low_attendance_staff,
            'peak_absence_days': peak_absence,
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def staff_attendance_filtered_view(request):
    """Date-based filtered attendance with search, department, and status filters."""
    try:
        date_param = request.query_params.get('date')
        search = request.query_params.get('search', '').strip().lower()
        department = request.query_params.get('department', '').strip()
        status_filter = request.query_params.get('status', '').strip()
        selected_date = datetime.datetime.strptime(date_param, '%Y-%m-%d').date() if date_param else timezone.now().date()

        att_qs = StaffAttendance.objects.filter(date=selected_date).select_related('staff__user', 'staff__user__role').order_by('staff__user__first_name')
        if department:
            att_qs = att_qs.filter(
                models.Q(staff__department__iexact=department) |
                models.Q(staff__user__role__name__iexact=department)
            )

        # Include unmarked staff
        all_staff = Staff.objects.select_related('user', 'user__role').filter(status='active')
        if department:
            all_staff = all_staff.filter(
                models.Q(department__iexact=department) |
                models.Q(user__role__name__iexact=department)
            )

        marked_staff_ids = set(att_qs.values_list('staff_id', flat=True))
        results = []
        for a in att_qs:
            s = a.staff
            if search and search not in (s.user.get_full_name() or s.user.username).lower():
                continue
            is_late_flag = bool(a.is_late or (a.late_minutes or 0) > 0 or a.status == 'late')
            ui_status = 'late' if is_late_flag else a.status

            if status_filter:
                if status_filter == 'unmarked':
                    continue
                if status_filter == 'late':
                    if not is_late_flag:
                        continue
                elif ui_status != status_filter:
                    continue
            results.append({
                'id': a.id,
                'staff_id': s.id,
                'staff_name': s.user.get_full_name() or s.user.username,
                'role': s.user.role.name if s.user.role else 'Unassigned',
                'department': _get_staff_department(s),
                'status': ui_status,
                'in_time': a.in_time.strftime('%H:%M') if a.in_time else None,
                'out_time': a.out_time.strftime('%H:%M') if a.out_time else None,
                'working_hours': float(a.working_hours) if a.working_hours else None,
                'late_minutes': a.late_minutes,
                'overtime': float(a.overtime) if a.overtime else 0,
                'is_late': is_late_flag,
                'remarks': a.remarks,
            })

        for s in all_staff:
            if s.id in marked_staff_ids:
                continue
            if search and search not in (s.user.get_full_name() or s.user.username).lower():
                continue
            if status_filter and status_filter != 'unmarked':
                continue
            results.append({
                'id': None,
                'staff_id': s.id,
                'staff_name': s.user.get_full_name() or s.user.username,
                'role': s.user.role.name if s.user.role else 'Unassigned',
                'department': _get_staff_department(s),
                'status': 'unmarked',
                'in_time': None,
                'out_time': None,
                'working_hours': None,
                'late_minutes': 0,
                'overtime': 0,
                'is_late': False,
                'remarks': '',
            })

        return Response(results, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def staff_bulk_attendance_view(request):
    """Bulk mark or update staff attendance for a date."""
    try:
        date = request.data.get('date')
        records = request.data.get('records', [])
        if not date:
            return Response({'error': 'date is required'}, status=status.HTTP_400_BAD_REQUEST)
        selected_date = datetime.datetime.strptime(date, '%Y-%m-%d').date()

        created, updated = 0, 0
        errors = []
        for rec in records:
            staff_id = rec.get('staff_id')
            att_status = (rec.get('status', 'present') or '').strip()
            force_late = att_status == 'late'
            if force_late:
                att_status = 'present'
            if att_status == 'unmarked':
                continue
            in_time = rec.get('in_time')
            out_time = rec.get('out_time')
            late_minutes = rec.get('late_minutes', 0)
            working_hours = rec.get('working_hours')
            overtime = rec.get('overtime', 0)
            remarks = rec.get('remarks', '')
            explicit_is_late = _to_bool(rec.get('is_late'))

            try:
                staff = Staff.objects.get(pk=staff_id)
            except Staff.DoesNotExist:
                errors.append({'staff_id': staff_id, 'error': 'Staff not found'})
                continue

            defaults = {
                'status': att_status,
                'marked_by': request.user,
                'remarks': remarks,
                'late_minutes': max(0, _to_int(late_minutes, 0)),
                'working_hours': working_hours,
                'overtime': overtime,
            }
            if in_time:
                try:
                    defaults['in_time'] = datetime.datetime.strptime(in_time, '%H:%M').time()
                except ValueError:
                    pass
            if out_time:
                try:
                    defaults['out_time'] = datetime.datetime.strptime(out_time, '%H:%M').time()
                except ValueError:
                    pass

            if att_status in ('absent', 'on_leave'):
                defaults['is_late'] = False
                defaults['late_minutes'] = 0
            elif explicit_is_late is not None:
                defaults['is_late'] = explicit_is_late
                if not explicit_is_late:
                    defaults['late_minutes'] = 0
            else:
                is_late, computed_minutes = _compute_late_fields(
                    selected_date=selected_date,
                    in_time=defaults.get('in_time'),
                    late_minutes=defaults.get('late_minutes', 0),
                    force_late=force_late,
                )
                defaults['is_late'] = is_late
                defaults['late_minutes'] = computed_minutes

            # Auto-calculate working hours if both times present
            if defaults.get('in_time') and defaults.get('out_time') and working_hours is None:
                delta = datetime.datetime.combine(selected_date, defaults['out_time']) - datetime.datetime.combine(selected_date, defaults['in_time'])
                defaults['working_hours'] = round(delta.total_seconds() / 3600, 2)

            obj, was_created = StaffAttendance.objects.update_or_create(
                staff=staff,
                date=selected_date,
                defaults=defaults,
            )
            if was_created:
                created += 1
            else:
                updated += 1

        return Response({'message': f'Saved. Created: {created}, Updated: {updated}.', 'errors': errors}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PATCH'])
@permission_classes([permissions.IsAuthenticated])
def staff_attendance_update_view(request, pk):
    """Inline edit a single attendance record."""
    try:
        att = StaffAttendance.objects.get(pk=pk)
    except StaffAttendance.DoesNotExist:
        return Response({'error': 'Attendance record not found'}, status=status.HTTP_404_NOT_FOUND)

    data = request.data
    status_value = data.get('status') if 'status' in data else None
    if status_value is not None:
        status_value = (str(status_value) or '').strip()
        if status_value == 'unmarked':
            att.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        if status_value == 'late':
            att.status = 'present'
        else:
            att.status = status_value
    if 'in_time' in data:
        try:
            att.in_time = datetime.datetime.strptime(data['in_time'], '%H:%M').time()
        except ValueError:
            pass
    if 'out_time' in data:
        try:
            att.out_time = datetime.datetime.strptime(data['out_time'], '%H:%M').time()
        except ValueError:
            pass
    if 'late_minutes' in data:
        att.late_minutes = max(0, _to_int(data['late_minutes'], 0))
    if 'working_hours' in data:
        att.working_hours = data['working_hours']
    if 'overtime' in data:
        att.overtime = data['overtime']
    if 'remarks' in data:
        att.remarks = data['remarks']
    if 'is_late' in data:
        parsed = _to_bool(data['is_late'])
        if parsed is not None:
            att.is_late = parsed
        else:
            att.is_late = bool(data['is_late'])

    # Keep `is_late` and `late_minutes` consistent when status/in_time/late_minutes changes.
    if att.status in ('absent', 'on_leave'):
        att.is_late = False
        att.late_minutes = 0
    else:
        force_late = (status_value == 'late') if status_value is not None else False
        if 'is_late' not in data:
            is_late, computed_minutes = _compute_late_fields(
                selected_date=att.date,
                in_time=att.in_time,
                late_minutes=att.late_minutes,
                force_late=force_late,
            )
            att.is_late = is_late
            att.late_minutes = computed_minutes
        elif not att.is_late:
            att.late_minutes = 0

    # Auto-calculate working hours if both times present and not explicitly set.
    if 'working_hours' not in data and att.in_time and att.out_time:
        delta = datetime.datetime.combine(att.date, att.out_time) - datetime.datetime.combine(att.date, att.in_time)
        att.working_hours = round(delta.total_seconds() / 3600, 2)

    att.save()
    return Response(StaffAttendanceSerializer(att).data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def staff_leave_balance_view(request):
    """Return leave balances for all staff or a specific staff member."""
    try:
        staff_id = request.query_params.get('staff')
        department = request.query_params.get('department')
        year = int(request.query_params.get('year', timezone.now().year))
        staff_qs = Staff.objects.select_related('user', 'user__role').filter(status='active')
        if staff_id:
            staff_qs = staff_qs.filter(pk=staff_id)
        if department:
            staff_qs = staff_qs.filter(
                models.Q(department__iexact=department) |
                models.Q(user__role__name__iexact=department)
            )

        results = []
        for s in staff_qs:
            balances = []
            for lb in s.leave_balances.filter(year=year):
                balances.append({
                    'leave_type': lb.leave_type,
                    'total_allowed': lb.total_allowed,
                    'used': lb.used,
                    'remaining': lb.remaining,
                })
            # If no balance records exist, create default ones from policy
            if not balances:
                school = getattr(request.user, 'school', None)
                if school:
                    policies = LeavePolicy.objects.filter(school=school)
                    for pol in policies:
                        balances.append({
                            'leave_type': pol.leave_type,
                            'total_allowed': pol.days_per_year,
                            'used': 0,
                            'remaining': pol.days_per_year,
                        })
            results.append({
                'staff_id': s.id,
                'staff_name': s.user.get_full_name() or s.user.username,
                'department': _get_staff_department(s),
                'balances': balances,
            })
        return Response(results, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def staff_monthly_report_view(request):
    """Generate monthly attendance summary for all staff."""
    try:
        month = int(request.query_params.get('month', timezone.now().month))
        year = int(request.query_params.get('year', timezone.now().year))
        department = request.query_params.get('department')
        staff_qs = Staff.objects.select_related('user', 'user__role').filter(status='active')
        staff_qs = _filter_staff_by_department(staff_qs, department)

        report = []
        for s in staff_qs:
            att_records = StaffAttendance.objects.filter(staff=s, date__month=month, date__year=year)
            total_working_days = att_records.count()
            days_present = att_records.filter(status='present', is_late=False).count()
            days_absent = att_records.filter(status='absent').count()
            days_half_day = att_records.filter(status='half_day').count()
            days_leave = att_records.filter(status='on_leave').count()
            days_late = att_records.filter(_late_attendance_q()).count()
            late_count = days_late
            total_late_minutes = sum(a.late_minutes or 0 for a in att_records.filter(_late_attendance_q()))
            total_ot = sum(float(a.overtime or 0) for a in att_records)
            total_working_hours = sum(float(a.working_hours or 0) for a in att_records if a.working_hours)

            leaves = StaffLeaveRequest.objects.filter(staff=s, status='approved', from_date__month=month, from_date__year=year)
            leave_by_type = {'casual': 0, 'sick': 0, 'earned': 0, 'emergency': 0}
            for l in leaves:
                delta = (l.to_date - l.from_date).days + 1
                leave_by_type[l.leave_type] = leave_by_type.get(l.leave_type, 0) + delta

            payable_days = days_present + days_late + (days_half_day * 0.5) + days_leave

            report.append({
                'staff_id': s.id,
                'staff_name': s.user.get_full_name() or s.user.username,
                'employee_id': s.employee_id,
                'department': _get_staff_department(s),
                'total_working_days': total_working_days,
                'days_present': days_present,
                'days_absent': days_absent,
                'days_half_day': days_half_day,
                'days_leave': days_leave,
                'leave_by_type': leave_by_type,
                'late_count': late_count,
                'total_late_minutes': total_late_minutes,
                'total_overtime': round(total_ot, 2),
                'total_working_hours': round(total_working_hours, 2),
                'payable_days': round(payable_days, 1),
            })
        return Response({'month': month, 'year': year, 'report': report}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def staff_monthly_report_export_view(request):
    """Export monthly attendance report as Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        month = int(request.query_params.get('month', timezone.now().month))
        year = int(request.query_params.get('year', timezone.now().year))
        department = request.query_params.get('department')
        month_name = datetime.date(year, month, 1).strftime('%B_%Y')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance Report'

        headers = ['Staff Name', 'Employee ID', 'Department', 'Total Working Days', 'Days Present', 'Days Absent',
                   'Days Half Day', 'Leaves Taken (Casual)', 'Leaves Taken (Sick)', 'Leaves Taken (Earned)',
                   'Late Count', 'Total Late Minutes', 'Total Overtime (hrs)', 'Total Working Hours', 'Payable Days']
        ws.append(headers)

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        staff_qs = Staff.objects.select_related('user', 'user__role').filter(status='active')
        staff_qs = _filter_staff_by_department(staff_qs, department)
        for s in staff_qs:
            att_records = StaffAttendance.objects.filter(staff=s, date__month=month, date__year=year)
            days_present = att_records.filter(status='present', is_late=False).count()
            days_absent = att_records.filter(status='absent').count()
            days_half_day = att_records.filter(status='half_day').count()
            days_leave = att_records.filter(status='on_leave').count()
            days_late = att_records.filter(_late_attendance_q()).count()
            late_count = days_late
            total_late_minutes = sum(a.late_minutes or 0 for a in att_records.filter(_late_attendance_q()))
            total_ot = sum(float(a.overtime or 0) for a in att_records)
            total_working_hours = sum(float(a.working_hours or 0) for a in att_records if a.working_hours)

            leaves = StaffLeaveRequest.objects.filter(staff=s, status='approved', from_date__month=month, from_date__year=year)
            leave_by_type = {'casual': 0, 'sick': 0, 'earned': 0}
            for l in leaves:
                delta = (l.to_date - l.from_date).days + 1
                if l.leave_type in leave_by_type:
                    leave_by_type[l.leave_type] = leave_by_type[l.leave_type] + delta

            payable_days = days_present + days_late + (days_half_day * 0.5) + days_leave
            ws.append([
                s.user.get_full_name() or s.user.username,
                s.employee_id,
                _get_staff_department(s),
                att_records.count(),
                days_present,
                days_absent,
                days_half_day,
                leave_by_type['casual'],
                leave_by_type['sick'],
                leave_by_type['earned'],
                late_count,
                total_late_minutes,
                round(total_ot, 2),
                round(total_working_hours, 2),
                round(payable_days, 1),
            ])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{month_name}.xlsx"'
        return response
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
