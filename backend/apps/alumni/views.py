import csv
import io
from datetime import timedelta

from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import permissions, status, viewsets, filters
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from .models import (
    Alumni,
    AlumniEvent,
    AlumniEventRSVP,
    AlumniCommunicationLog,
    AlumniContribution,
    AlumniAchievement,
)
from .serializers import (
    AlumniSerializer,
    AlumniEventSerializer,
    AlumniEventRSVPSerializer,
    AlumniCommunicationLogSerializer,
    AlumniContributionSerializer,
    AlumniAchievementSerializer,
)


def _scope_to_school(qs, user, field="school_id"):
    if user.is_superuser:
        return qs
    school_id = getattr(user, "school_id", None)
    if school_id:
        return qs.filter(**{field: school_id})
    return qs.none()


class AlumniViewSet(viewsets.ModelViewSet):
    serializer_class = AlumniSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ["name", "email", "phone", "organization", "job_role", "industry", "location"]

    def get_queryset(self):
        qs = Alumni.objects.all()
        qs = _scope_to_school(qs, self.request.user)

        params = self.request.query_params
        year = params.get("graduation_year") or params.get("year")
        industry = params.get("industry")
        location = params.get("location")
        status_param = params.get("status")
        is_verified = params.get("is_verified")

        if year and year.isdigit():
            qs = qs.filter(graduation_year=int(year))
        if industry:
            qs = qs.filter(industry__icontains=industry.strip())
        if location:
            qs = qs.filter(location__icontains=location.strip())
        if status_param:
            qs = qs.filter(status=status_param)
        if is_verified is not None and is_verified != "":
            if str(is_verified).lower() in ("true", "1", "yes"):
                qs = qs.filter(is_verified=True)
            elif str(is_verified).lower() in ("false", "0", "no"):
                qs = qs.filter(is_verified=False)

        return qs

    def perform_create(self, serializer):
        data = serializer.validated_data
        user = self.request.user
        if not user.is_superuser and getattr(user, "school_id", None) and not data.get("school"):
            serializer.save(school_id=user.school_id, created_by=user)
            return
        serializer.save(created_by=user)

    def perform_update(self, serializer):
        serializer.save()

    @action(detail=False, methods=["get"], url_path="dashboard")
    def dashboard(self, request):
        alumni_qs = self.get_queryset()

        total = alumni_qs.exclude(status="rejected").count()
        recently_added = alumni_qs.filter(created_at__gte=timezone.now() - timedelta(days=7)).count()
        pending = alumni_qs.filter(status="pending").count()

        events_qs = _scope_to_school(AlumniEvent.objects.all(), request.user)
        upcoming_events = events_qs.filter(start_at__gte=timezone.now()).count()

        return Response(
            {
                "total_alumni_count": total,
                "recently_added_alumni": recently_added,
                "pending_approvals": pending,
                "upcoming_events": upcoming_events,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        alumni = self.get_object()
        alumni.status = "approved"
        alumni.save(update_fields=["status"])
        return Response({"message": "Alumni approved."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="reject")
    def reject(self, request, pk=None):
        alumni = self.get_object()
        alumni.status = "rejected"
        alumni.save(update_fields=["status"])
        return Response({"message": "Alumni rejected."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="verify")
    def verify(self, request, pk=None):
        alumni = self.get_object()
        alumni.mark_verified(user=request.user)
        alumni.save(update_fields=["is_verified", "verified_at", "verified_by"])
        return Response({"message": "Alumni verified."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="unverify")
    def unverify(self, request, pk=None):
        alumni = self.get_object()
        alumni.is_verified = False
        alumni.verified_at = None
        alumni.verified_by = None
        alumni.save(update_fields=["is_verified", "verified_at", "verified_by"])
        return Response({"message": "Alumni marked unverified."}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="import", parser_classes=[MultiPartParser, FormParser])
    def bulk_import(self, request):
        """
        Bulk upload via CSV/XLSX.
        Expected columns (case-insensitive):
          - name, graduation_year/year, class_stream, email, phone, job_role, organization, industry, location
        """
        upload = request.FILES.get("file")
        if not upload:
            return Response({"error": "file is required"}, status=status.HTTP_400_BAD_REQUEST)

        default_status = request.data.get("default_status") or "pending"
        if default_status not in {"pending", "approved", "rejected"}:
            return Response({"error": "default_status must be pending/approved/rejected"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            filename = (upload.name or "").lower()
            if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
                rows = self._read_xlsx(upload)
            else:
                rows = self._read_csv(upload)
        except Exception as e:
            return Response({"error": f"Failed to parse file: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        user = request.user
        import_school_id = None
        if user.is_superuser:
            import_school_id = request.data.get("school") or request.data.get("school_id") or getattr(user, "school_id", None)
        else:
            import_school_id = getattr(user, "school_id", None)

        if not import_school_id:
            return Response({"error": "school_id is required for import."}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        errors = []

        with transaction.atomic():
            for idx, row in enumerate(rows, start=1):
                try:
                    payload = self._normalize_import_row(row)
                    if not payload.get("name") or not payload.get("graduation_year"):
                        raise ValueError("name and graduation_year are required")

                    Alumni.objects.create(
                        school_id=int(import_school_id),
                        created_by=user,
                        status=payload.get("status") or default_status,
                        is_verified=bool(payload.get("is_verified") or False),
                        name=payload["name"],
                        graduation_year=payload["graduation_year"],
                        class_stream=payload.get("class_stream", ""),
                        email=payload.get("email", ""),
                        phone=payload.get("phone", ""),
                        job_role=payload.get("job_role", ""),
                        organization=payload.get("organization", ""),
                        industry=payload.get("industry", ""),
                        location=payload.get("location", ""),
                    )
                    created += 1
                except Exception as e:
                    errors.append({"row": idx, "error": str(e)})

        return Response({"created": created, "errors": errors}, status=status.HTTP_200_OK)

    def _read_csv(self, upload):
        content = upload.read()
        # Try UTF-8, fallback to latin-1
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        buf = io.StringIO(text)
        reader = csv.DictReader(buf)
        return list(reader)

    def _read_xlsx(self, upload):
        from openpyxl import load_workbook

        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        out = []
        for r in rows[1:]:
            row = {}
            for i, h in enumerate(headers):
                row[h] = r[i] if i < len(r) else None
            out.append(row)
        return out

    def _normalize_import_row(self, row):
        # Case-insensitive keys with a few aliases
        normalized = {}
        for k, v in (row or {}).items():
            key = str(k or "").strip().lower()
            if not key:
                continue
            normalized[key] = v

        def get_any(*keys):
            for k in keys:
                if k in normalized and normalized[k] not in (None, ""):
                    return normalized[k]
            return None

        graduation_year = get_any("graduation_year", "year", "batch", "batch_year")
        if graduation_year not in (None, ""):
            try:
                graduation_year = int(str(graduation_year).strip())
            except Exception:
                graduation_year = None

        status_val = get_any("status")
        if status_val is not None:
            status_val = str(status_val).strip().lower()

        is_verified = get_any("is_verified", "verified")
        if is_verified is not None:
            is_verified = str(is_verified).strip().lower() in ("true", "1", "yes", "y")

        return {
            "name": str(get_any("name") or "").strip(),
            "graduation_year": graduation_year,
            "class_stream": str(get_any("class_stream", "class", "stream") or "").strip(),
            "email": str(get_any("email") or "").strip(),
            "phone": str(get_any("phone", "mobile") or "").strip(),
            "job_role": str(get_any("job_role", "role") or "").strip(),
            "organization": str(get_any("organization", "company") or "").strip(),
            "industry": str(get_any("industry") or "").strip(),
            "location": str(get_any("location", "city") or "").strip(),
            "status": status_val if status_val in {"pending", "approved", "rejected"} else None,
            "is_verified": is_verified,
        }

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        qs = self.get_queryset().order_by("graduation_year", "name")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="alumni_export.csv"'

        writer = csv.writer(response)
        writer.writerow(
            [
                "id",
                "name",
                "graduation_year",
                "class_stream",
                "email",
                "phone",
                "job_role",
                "organization",
                "industry",
                "location",
                "status",
                "is_verified",
                "created_at",
            ]
        )
        for a in qs:
            writer.writerow(
                [
                    a.id,
                    a.name,
                    a.graduation_year,
                    a.class_stream,
                    a.email,
                    a.phone,
                    a.job_role,
                    a.organization,
                    a.industry,
                    a.location,
                    a.status,
                    a.is_verified,
                    a.created_at.isoformat() if a.created_at else "",
                ]
            )
        return response

    @action(detail=False, methods=["get"], url_path="reports")
    def reports(self, request):
        qs = self.get_queryset().exclude(status="rejected")

        by_year = (
            qs.values("graduation_year")
            .annotate(count=Count("id"))
            .order_by("graduation_year")
        )

        career_dist = (
            qs.exclude(industry="")
            .values("industry")
            .annotate(count=Count("id"))
            .order_by("-count")[:20]
        )

        verified_counts = qs.aggregate(
            verified=Count("id", filter=Q(is_verified=True)),
            unverified=Count("id", filter=Q(is_verified=False)),
        )

        return Response(
            {
                "alumni_by_year": list(by_year),
                "career_distribution": list(career_dist),
                "verification": verified_counts,
            },
            status=status.HTTP_200_OK,
        )


class AlumniEventViewSet(viewsets.ModelViewSet):
    serializer_class = AlumniEventSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ["title", "description", "venue"]

    def get_queryset(self):
        qs = AlumniEvent.objects.all()
        qs = _scope_to_school(qs, self.request.user)

        params = self.request.query_params
        upcoming = params.get("upcoming")
        if upcoming and str(upcoming).lower() in ("true", "1", "yes"):
            qs = qs.filter(start_at__gte=timezone.now())
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        data = serializer.validated_data
        if not user.is_superuser and getattr(user, "school_id", None) and not data.get("school"):
            serializer.save(school_id=user.school_id, created_by=user)
            return
        serializer.save(created_by=user)

    @action(detail=True, methods=["get"], url_path="rsvps")
    def rsvps(self, request, pk=None):
        event = self.get_object()
        qs = event.rsvps.select_related("alumni").all()
        serializer = AlumniEventRSVPSerializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AlumniEventRSVPViewSet(viewsets.ModelViewSet):
    serializer_class = AlumniEventRSVPSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = AlumniEventRSVP.objects.select_related("event", "alumni").all()
        qs = _scope_to_school(qs, self.request.user, field="event__school_id")
        event_id = self.request.query_params.get("event")
        if event_id and str(event_id).isdigit():
            qs = qs.filter(event_id=int(event_id))
        return qs

    def perform_create(self, serializer):
        serializer.save()


class AlumniCommunicationLogViewSet(viewsets.ModelViewSet):
    serializer_class = AlumniCommunicationLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = AlumniCommunicationLog.objects.all()
        return _scope_to_school(qs, self.request.user)

    def perform_create(self, serializer):
        user = self.request.user
        data = serializer.validated_data
        if not user.is_superuser and getattr(user, "school_id", None) and not data.get("school"):
            serializer.save(school_id=user.school_id, created_by=user)
            return
        serializer.save(created_by=user)

    @action(detail=True, methods=["post"], url_path="send")
    def send(self, request, pk=None):
        """
        Records a send attempt and marks log as 'sent'.
        Actual delivery is intentionally out-of-scope (can be wired to Celery later).
        """
        log = self.get_object()

        # Compute recipient count based on segment filters (best-effort).
        filters_in = log.segment_filters or {}
        alumni_qs = _scope_to_school(Alumni.objects.exclude(status="rejected"), request.user)
        alumni_qs = alumni_qs.filter(status="approved")

        year = filters_in.get("graduation_year") or filters_in.get("year")
        location = filters_in.get("location")
        industry = filters_in.get("industry")
        is_verified = filters_in.get("is_verified")

        if year and str(year).isdigit():
            alumni_qs = alumni_qs.filter(graduation_year=int(year))
        if location:
            alumni_qs = alumni_qs.filter(location__icontains=str(location))
        if industry:
            alumni_qs = alumni_qs.filter(industry__icontains=str(industry))
        if is_verified is not None and str(is_verified) != "":
            if str(is_verified).lower() in ("true", "1", "yes"):
                alumni_qs = alumni_qs.filter(is_verified=True)
            elif str(is_verified).lower() in ("false", "0", "no"):
                alumni_qs = alumni_qs.filter(is_verified=False)

        recipient_count = alumni_qs.count()

        log.status = "sent"
        log.recipient_count = recipient_count
        log.sent_at = timezone.now()
        log.save(update_fields=["status", "recipient_count", "sent_at"])

        return Response(
            {"message": "Marked as sent.", "recipient_count": recipient_count},
            status=status.HTTP_200_OK,
        )


class AlumniContributionViewSet(viewsets.ModelViewSet):
    serializer_class = AlumniContributionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = AlumniContribution.objects.select_related("alumni").all()
        qs = _scope_to_school(qs, self.request.user)
        alumni_id = self.request.query_params.get("alumni")
        if alumni_id and str(alumni_id).isdigit():
            qs = qs.filter(alumni_id=int(alumni_id))
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        data = serializer.validated_data
        if not user.is_superuser and getattr(user, "school_id", None) and not data.get("school"):
            serializer.save(school_id=user.school_id, created_by=user)
            return
        serializer.save(created_by=user)


class AlumniAchievementViewSet(viewsets.ModelViewSet):
    serializer_class = AlumniAchievementSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = AlumniAchievement.objects.select_related("alumni").all()
        qs = _scope_to_school(qs, self.request.user)
        visibility = self.request.query_params.get("visibility")
        featured = self.request.query_params.get("is_featured")

        if visibility:
            qs = qs.filter(visibility=visibility)
        if featured is not None and str(featured) != "":
            if str(featured).lower() in ("true", "1", "yes"):
                qs = qs.filter(is_featured=True)
            elif str(featured).lower() in ("false", "0", "no"):
                qs = qs.filter(is_featured=False)
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        data = serializer.validated_data
        if not user.is_superuser and getattr(user, "school_id", None) and not data.get("school"):
            serializer.save(school_id=user.school_id, created_by=user)
            return
        serializer.save(created_by=user)
