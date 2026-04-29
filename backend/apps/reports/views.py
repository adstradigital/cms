import io
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse

from .models import SavedReport
from .serializers import SavedReportSerializer
from .query_engine import MODULE_META, run_report, run_report_full


# ── Meta endpoint ─────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def report_meta(request):
    """Return available modules with their field definitions."""
    return Response(MODULE_META)


# ── Run report ────────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def run_report_view(request):
    data = request.data
    module = data.get('module', 'fees')
    fields = data.get('fields', [])
    filters = data.get('filters', [])
    filter_logic = data.get('filter_logic', 'AND')
    group_by = data.get('group_by', '')
    sort_field = data.get('sort_field', '')
    sort_dir = data.get('sort_direction', 'desc')
    page = int(data.get('page', 1))
    page_size = min(int(data.get('page_size', 20)), 500)

    school = getattr(request.user, 'school', None)

    try:
        result = run_report(
            module=module,
            fields=fields,
            filters=filters,
            filter_logic=filter_logic,
            group_by=group_by,
            sort_field=sort_field,
            sort_dir=sort_dir,
            page=page,
            page_size=page_size,
            school=school,
        )
        return Response(result)
    except ValueError as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': f'Query failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ── Export Excel ──────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.data
    module = data.get('module', 'fees')
    fields = data.get('fields', [])
    filters = data.get('filters', [])
    filter_logic = data.get('filter_logic', 'AND')
    group_by = data.get('group_by', '')
    sort_field = data.get('sort_field', '')
    sort_dir = data.get('sort_direction', 'desc')
    report_name = data.get('report_name', 'Report')
    school = getattr(request.user, 'school', None)

    try:
        result = run_report_full(module, fields, filters, filter_logic,
                                  group_by, sort_field, sort_dir, school)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    module_fields = MODULE_META.get(module, {}).get('fields', [])
    selected_field_defs = [f for f in module_fields if not fields or f['key'] in fields]
    if not selected_field_defs:
        selected_field_defs = module_fields

    wb = Workbook()
    ws = wb.active
    ws.title = report_name[:31]

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Title row
    ws.merge_cells(f'A1:{get_column_letter(len(selected_field_defs))}1')
    title_cell = ws['A1']
    title_cell.value = report_name
    title_cell.font = Font(bold=True, size=13, color='1E293B')
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    # Header row (row 2)
    for col_idx, field in enumerate(selected_field_defs, 1):
        cell = ws.cell(row=2, column=col_idx, value=field['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 22

    # Data rows
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    for row_idx, row_data in enumerate(result['rows'], 3):
        for col_idx, field in enumerate(selected_field_defs, 1):
            val = row_data.get(field['key'], '')
            if isinstance(val, bool):
                val = 'Yes' if val else 'No'
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 18

    # Auto column widths
    for col_idx, field in enumerate(selected_field_defs, 1):
        max_len = max(
            len(field['label']),
            max((len(str(row.get(field['key'], '') or '')) for row in result['rows']), default=0)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{report_name}.xlsx"'
    return response


# ── Export PDF ────────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    data = request.data
    module = data.get('module', 'fees')
    fields = data.get('fields', [])
    filters = data.get('filters', [])
    filter_logic = data.get('filter_logic', 'AND')
    group_by = data.get('group_by', '')
    sort_field = data.get('sort_field', '')
    sort_dir = data.get('sort_direction', 'desc')
    report_name = data.get('report_name', 'Report')
    school = getattr(request.user, 'school', None)

    try:
        result = run_report_full(module, fields, filters, filter_logic,
                                  group_by, sort_field, sort_dir, school)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    module_fields = MODULE_META.get(module, {}).get('fields', [])
    selected_field_defs = [f for f in module_fields if not fields or f['key'] in fields]
    if not selected_field_defs:
        selected_field_defs = module_fields

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                  fontSize=16, textColor=colors.HexColor('#1E293B'),
                                  spaceAfter=6)
    meta_style = ParagraphStyle('meta', parent=styles['Normal'],
                                 fontSize=8, textColor=colors.HexColor('#64748B'),
                                 spaceAfter=10)

    from datetime import date
    elements = [
        Paragraph(report_name, title_style),
        Paragraph(f"Module: {module.title()} | Records: {result['total']} | Generated: {date.today()}", meta_style),
        Spacer(1, 4*mm),
    ]

    header = [field['label'] for field in selected_field_defs]
    rows_data = [header]
    for row in result['rows']:
        r = []
        for field in selected_field_defs:
            v = row.get(field['key'], '')
            if isinstance(v, bool):
                v = 'Yes' if v else 'No'
            r.append(str(v) if v is not None else '')
        rows_data.append(r)

    col_count = len(selected_field_defs)
    available_width = landscape(A4)[0] - 30*mm
    col_width = available_width / col_count

    table = Table(rows_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_name}.pdf"'
    return response


# ── Saved Reports CRUD ────────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def saved_reports(request):
    if request.method == 'GET':
        qs = SavedReport.objects.filter(created_by=request.user)
        search = request.query_params.get('search', '')
        if search:
            qs = qs.filter(name__icontains=search)
        module = request.query_params.get('module', '')
        if module:
            qs = qs.filter(module=module)
        serializer = SavedReportSerializer(qs, many=True)
        return Response(serializer.data)

    # POST — create
    serializer = SavedReportSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(created_by=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def saved_report_detail(request, pk):
    try:
        report = SavedReport.objects.get(pk=pk, created_by=request.user)
    except SavedReport.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        return Response(SavedReportSerializer(report).data)

    if request.method in ('PUT', 'PATCH'):
        serializer = SavedReportSerializer(report, data=request.data,
                                            partial=(request.method == 'PATCH'))
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'DELETE':
        report.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
